import pandas as pd
# from icecream import ic
from devdriven.tempfile import tempfile_to_writeable, tempfile_from_readable
from .command import section, command
from .formats import FormatIn, FormatOut

section('Format', 20)

@command
class XlsIn(FormatIn):
  r'''
  xls-in - Read XLS Spreadsheet.
  alias: -xls

  --sheet-name=NAME  |  Sheet name
  --header, -h       |  Generate header.  Default: True.

  :suffixes: .xlsx

  Examples:

$ psv in a.xlsx // -xls // md

$ psv in a.xlsx // -xls --no-header // md

  '''
  def format_in(self, readable, _env):
    # pylint: disable-next=import-outside-toplevel
    from openpyxl import load_workbook

    def read_workbook(filename):
      return load_workbook(filename=filename)

    workbook = tempfile_from_readable(readable, '.xlsx', read_workbook)
    sheet_id = self.opt('sheet-name', 0)
    worksheet = workbook.worksheets[sheet_id]
    data = worksheet.values
    if self.opt('header', True):
      cols = list(next(data))
      data = list(data)
    else:
      data = list(data)
      cols = [f'c{i}' for i in range(0, len(data[0]))]
    return pd.DataFrame(data, columns=cols)

  def wants_input_file(self):
    return False

  def default_encoding(self):
    return None

@command
class XlsOut(FormatOut):
  '''
  xls-out - Generate XLS Spreadsheet.
  alias: xls-,xls

  --sheet-name=NAME  |  Sheet name
  --header, -h       |  Generate header.  Default: True.

  :suffix=.xlsx

  Examples:

$ psv in a.csv // xls // o a.xlsx
$ file a.xlsx

  '''
  def format_out(self, inp, _env, writeable):
    # pylint: disable-next=import-outside-toplevel
    from openpyxl import Workbook
    # pylint: disable-next=import-outside-toplevel
    from openpyxl.utils.dataframe import dataframe_to_rows
    header = bool(self.opt('header', True))
    index = bool(self.opt('index', False))
    workbook = Workbook()
    worksheet = workbook.active

    def save_workbook(tmp_file):
      workbook.save(tmp_file)

    if isinstance(inp, pd.DataFrame):
      for row in dataframe_to_rows(inp, index=index, header=header):
        worksheet.append(row)
      tempfile_to_writeable(writeable, '.xlsx', save_workbook)
    else:
      raise Exception("xls-out: cannot format {type(inp)}")

  def wants_output_file(self):
    return False

  def default_encoding(self):
    return None
